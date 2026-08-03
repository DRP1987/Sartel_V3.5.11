"""Installation sheet dialog for filling and exporting installation data."""

import json
import os
import re
import sys
import shutil
import tempfile
from datetime import date
from typing import Any, Dict, List, Optional

from PyQt5.QtWidgets import (
    QCheckBox, QComboBox, QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QScrollArea, QWidget, QPushButton, QLabel, QLineEdit, QTextEdit,
    QFileDialog, QMessageBox, QGroupBox, QGridLayout, QSizePolicy, QFrame,
    QDialogButtonBox, QRadioButton,
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap

# Available CAN baudrates
_CAN_BAUDRATES = [125000, 250000, 500000, 1000000]


# ---------------------------------------------------------------------------
# Helpers: locate the config folder
# ---------------------------------------------------------------------------

def _config_dir() -> str:
    """Return the absolute path to the config directory."""
    candidates = [
        os.path.join(os.path.dirname(__file__), "..", "config"),
        os.path.join(getattr(sys, "_MEIPASS", ""), "config"),
    ]
    for p in candidates:
        if os.path.isdir(p):
            return os.path.abspath(p)
    # fallback – sibling of gui package
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "config"))


def _sheets_config_path() -> str:
    """Return the path to installation_sheets.json inside config/."""
    return os.path.join(_config_dir(), "installation_sheets.json")


def _load_sheets_config() -> List[Dict[str, Any]]:
    """Load and return the list of sheet definitions from installation_sheets.json."""
    path = _sheets_config_path()
    if not os.path.isfile(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data.get("sheets", [])
    except Exception:
        return []


def _excel_template_path(filename: str) -> str:
    """Return the absolute path for a template Excel file stored in config/."""
    return os.path.join(_config_dir(), filename)


def _load_config_names() -> List[str]:
    """Return the list of configuration names from configurations.json."""
    path = os.path.join(_config_dir(), "configurations.json")
    if not os.path.isfile(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return [c.get("name", "") for c in data.get("configurations", []) if c.get("name")]
    except Exception:
        return []


def _resolve_cell_ref(ws, cell_ref: str) -> str:
    """Resolve a cell reference to a writable top-left address.

    Handles two cases:
    - Range notation: ``"C10:E12"`` → returns ``"C10"`` (top-left of the range).
    - Single cell that belongs to a merged region in *ws*: returns the top-left
      cell of that merged region so openpyxl writes the value correctly.
    - Single cell outside any merge: returns *cell_ref* unchanged.
    """
    from openpyxl.utils import get_column_letter, column_index_from_string

    # Handle explicit range notation e.g. "C10:E12" or "c10:e12"
    cell_ref = cell_ref.strip()
    if ":" in cell_ref:
        top_left = cell_ref.split(":")[0].strip().upper()
        return top_left

    # Parse the column/row indices of the single cell
    m = re.match(r"([A-Za-z]+)(\d+)$", cell_ref)
    if not m:
        return cell_ref  # unrecognised format – pass through unchanged

    col_letter, row_str = m.groups()
    col_idx = column_index_from_string(col_letter.upper())
    row_idx = int(row_str)

    # Check every merged region in the worksheet
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row_idx <= merged_range.max_row
            and merged_range.min_col <= col_idx <= merged_range.max_col
        ):
            return f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"

    return cell_ref


def _com_cell_ref(cell_ref: str) -> str:
    """Return the cell address to use with win32com ``Range()``.

    For range notation such as ``"C10:E12"`` the top-left address (``"C10"``)
    is returned so that only the first cell in a merged region is written.
    Single-cell references are returned unchanged (uppercased).
    """
    cell_ref = cell_ref.strip().upper()
    if ":" in cell_ref:
        return cell_ref.split(":")[0].strip()
    return cell_ref


# ---------------------------------------------------------------------------
# Conditional (Yes/No) widget
# ---------------------------------------------------------------------------

class ConditionalWidget(QWidget):
    """A compound widget for yes/no conditional questions.

    Renders two radio buttons ("Yes" / "No").  When the user selects "Yes"
    up to two configurable text boxes appear; when "No" is selected up to two
    different text boxes appear.  The boolean answer is written to
    *checkbox_cell* in the Excel template; the visible text content is written
    to *yes_text_cell* / *yes_text_cell2* or *no_text_cell* / *no_text_cell2*
    depending on the selection.

    JSON field definition example::

        {
          "label": "Is the crane control CanBus connected for LMB and Crane operation?",
          "type": "conditional",
          "checkbox_cell": "C34",
          "yes_text_label": "Connection Details",
          "yes_text_cell": "D34",
          "yes_text_label2": "Additional Yes Notes",
          "yes_text_cell2": "D35",
          "no_text_label": "Reason for Non-Connection",
          "no_text_cell": "E34",
          "no_text_label2": "Additional No Notes",
          "no_text_cell2": "E35"
        }
    """

    def __init__(self, field_def: Dict[str, Any], parent=None):
        super().__init__(parent)
        self._checkbox_cell: str = field_def.get("checkbox_cell", "")
        self._yes_text_cell: str = field_def.get("yes_text_cell", "")
        self._yes_text_cell2: str = field_def.get("yes_text_cell2", "")
        self._no_text_cell: str = field_def.get("no_text_cell", "")
        self._no_text_cell2: str = field_def.get("no_text_cell2", "")
        yes_label = field_def.get("yes_text_label", "Details (Yes)")
        yes_label2 = field_def.get("yes_text_label2", "")
        no_label = field_def.get("no_text_label", "Details (No)")
        no_label2 = field_def.get("no_text_label2", "")

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        # --- Yes / No radio buttons ---
        radio_row = QHBoxLayout()
        radio_row.setSpacing(16)
        self._yes_radio = QRadioButton("Yes")
        self._no_radio = QRadioButton("No")
        self._yes_radio.setStyleSheet("font-size: 9pt;")
        self._no_radio.setStyleSheet("font-size: 9pt;")
        radio_row.addWidget(self._yes_radio)
        radio_row.addWidget(self._no_radio)
        radio_row.addStretch()
        layout.addLayout(radio_row)

        # --- Text boxes shown when "Yes" is selected ---
        self._yes_container: Optional[QWidget] = None
        self._yes_text: Optional[QTextEdit] = None
        self._yes_text2: Optional[QTextEdit] = None
        if self._yes_text_cell:
            self._yes_container = QWidget()
            yes_inner = QVBoxLayout()
            yes_inner.setContentsMargins(0, 2, 0, 0)
            yes_inner.setSpacing(2)
            yes_lbl = QLabel(f"{yes_label}:")
            yes_lbl.setStyleSheet("font-size: 8pt; color: #1a6a2e; font-weight: bold;")
            self._yes_text = QTextEdit()
            self._yes_text.setPlaceholderText(yes_label)
            self._yes_text.setFixedHeight(60)
            self._yes_text.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            yes_inner.addWidget(yes_lbl)
            yes_inner.addWidget(self._yes_text)
            # Optional second text box for "Yes"
            if self._yes_text_cell2:
                lbl2_text = yes_label2 if yes_label2 else f"{yes_label} (2)"
                yes_lbl2 = QLabel(f"{lbl2_text}:")
                yes_lbl2.setStyleSheet("font-size: 8pt; color: #1a6a2e; font-weight: bold;")
                self._yes_text2 = QTextEdit()
                self._yes_text2.setPlaceholderText(lbl2_text)
                self._yes_text2.setFixedHeight(60)
                self._yes_text2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                yes_inner.addWidget(yes_lbl2)
                yes_inner.addWidget(self._yes_text2)
            self._yes_container.setLayout(yes_inner)
            self._yes_container.setVisible(False)
            layout.addWidget(self._yes_container)

        # --- Text boxes shown when "No" is selected ---
        self._no_container: Optional[QWidget] = None
        self._no_text: Optional[QTextEdit] = None
        self._no_text2: Optional[QTextEdit] = None
        if self._no_text_cell:
            self._no_container = QWidget()
            no_inner = QVBoxLayout()
            no_inner.setContentsMargins(0, 2, 0, 0)
            no_inner.setSpacing(2)
            no_lbl = QLabel(f"{no_label}:")
            no_lbl.setStyleSheet("font-size: 8pt; color: #a93226; font-weight: bold;")
            self._no_text = QTextEdit()
            self._no_text.setPlaceholderText(no_label)
            self._no_text.setFixedHeight(60)
            self._no_text.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            no_inner.addWidget(no_lbl)
            no_inner.addWidget(self._no_text)
            # Optional second text box for "No"
            if self._no_text_cell2:
                lbl2_text = no_label2 if no_label2 else f"{no_label} (2)"
                no_lbl2 = QLabel(f"{lbl2_text}:")
                no_lbl2.setStyleSheet("font-size: 8pt; color: #a93226; font-weight: bold;")
                self._no_text2 = QTextEdit()
                self._no_text2.setPlaceholderText(lbl2_text)
                self._no_text2.setFixedHeight(60)
                self._no_text2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                no_inner.addWidget(no_lbl2)
                no_inner.addWidget(self._no_text2)
            self._no_container.setLayout(no_inner)
            self._no_container.setVisible(False)
            layout.addWidget(self._no_container)

        self.setLayout(layout)

        # Connect radio button signals
        self._yes_radio.toggled.connect(self._on_selection_changed)
        self._no_radio.toggled.connect(self._on_selection_changed)

    # ------------------------------------------------------------------
    # Internal slot
    # ------------------------------------------------------------------

    def _on_selection_changed(self):
        """Show/hide the appropriate text boxes when the selection changes."""
        yes_checked = self._yes_radio.isChecked()
        no_checked = self._no_radio.isChecked()
        if self._yes_container is not None:
            self._yes_container.setVisible(yes_checked)
        if self._no_container is not None:
            self._no_container.setVisible(no_checked)

    # ------------------------------------------------------------------
    # Public helpers
    # ------------------------------------------------------------------

    def get_values(self) -> Dict[str, Any]:
        """Return a ``{cell_ref: value}`` mapping for all configured cells.

        * *checkbox_cell* → ``True`` (Yes), ``False`` (No), or omitted if
          neither radio button has been selected.
        * *yes_text_cell* / *yes_text_cell2* → text content when "Yes" is selected.
        * *no_text_cell* / *no_text_cell2*   → text content when "No" is selected.
        """
        result: Dict[str, Any] = {}

        if self._yes_radio.isChecked():
            is_yes: Optional[bool] = True
        elif self._no_radio.isChecked():
            is_yes = False
        else:
            is_yes = None

        if self._checkbox_cell and is_yes is not None:
            result[self._checkbox_cell] = is_yes

        if is_yes is True:
            if self._yes_text_cell and self._yes_text is not None:
                result[self._yes_text_cell] = self._yes_text.toPlainText().strip()
            if self._yes_text_cell2 and self._yes_text2 is not None:
                result[self._yes_text_cell2] = self._yes_text2.toPlainText().strip()
        elif is_yes is False:
            if self._no_text_cell and self._no_text is not None:
                result[self._no_text_cell] = self._no_text.toPlainText().strip()
            if self._no_text_cell2 and self._no_text2 is not None:
                result[self._no_text_cell2] = self._no_text2.toPlainText().strip()

        return result

    def clear(self):
        """Reset both radio buttons and clear all text boxes."""
        # Temporarily disable mutual exclusion to allow unchecking both
        self._yes_radio.setAutoExclusive(False)
        self._no_radio.setAutoExclusive(False)
        self._yes_radio.setChecked(False)
        self._no_radio.setChecked(False)
        self._yes_radio.setAutoExclusive(True)
        self._no_radio.setAutoExclusive(True)
        if self._yes_text is not None:
            self._yes_text.clear()
        if self._yes_text2 is not None:
            self._yes_text2.clear()
        if self._no_text is not None:
            self._no_text.clear()
        if self._no_text2 is not None:
            self._no_text2.clear()
        # Ensure both containers are hidden
        if self._yes_container is not None:
            self._yes_container.setVisible(False)
        if self._no_container is not None:
            self._no_container.setVisible(False)

    def set_values(self, values: Dict[str, Any]):
        """Populate widget from a ``{cell_ref: value}`` mapping (e.g. from a saved form)."""
        self.clear()
        bool_val = values.get(self._checkbox_cell)
        if bool_val is True:
            self._yes_radio.setChecked(True)
            if self._yes_text_cell and self._yes_text is not None:
                self._yes_text.setPlainText(str(values.get(self._yes_text_cell, "")))
            if self._yes_text_cell2 and self._yes_text2 is not None:
                self._yes_text2.setPlainText(str(values.get(self._yes_text_cell2, "")))
        elif bool_val is False:
            self._no_radio.setChecked(True)
            if self._no_text_cell and self._no_text is not None:
                self._no_text.setPlainText(str(values.get(self._no_text_cell, "")))
            if self._no_text_cell2 and self._no_text2 is not None:
                self._no_text2.setPlainText(str(values.get(self._no_text_cell2, "")))


# ---------------------------------------------------------------------------
# Checkbox-group widget
# ---------------------------------------------------------------------------

class CheckboxGroupWidget(QWidget):
    """A compound widget that renders a group of independent checkboxes.

    Each checkbox is linked to a specific Excel cell.  When a checkbox is
    ticked, an optional batch of sub-questions (any supported non-group field
    types) is revealed below it.  All sub-question widgets are hidden again
    when the checkbox is unchecked.

    JSON field definition example::

        {
          "label": "Installation Checks",
          "type": "checkbox_group",
          "group_id": "install_checks",
          "options": [
            {
              "label": "GPS Antenna Installed",
              "cell": "C30",
              "sub_questions": [
                {"label": "Antenna Location",   "cell": "D30", "type": "text"},
                {"label": "Signal Strength dB",  "cell": "E30", "type": "text"}
              ]
            },
            {
              "label": "CAN Bus Connected",
              "cell": "C31",
              "sub_questions": [
                {"label": "Connection Type",     "cell": "D31", "type": "text"},
                {"label": "Additional Notes",    "cell": "E31", "type": "multiline"}
              ]
            },
            {
              "label": "Power Supply Verified",
              "cell": "C32"
            },
            {
              "label": "Device Tested Online",
              "cell": "C33"
            }
          ]
        }
    """

    def __init__(self, field_def: Dict[str, Any], parent=None):
        super().__init__(parent)
        self._option_checkboxes: List[Dict[str, Any]] = []
        # Maps cell_ref → QWidget for sub-questions
        self._sub_widgets: Dict[str, QWidget] = {}

        outer = QVBoxLayout()
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(6)

        for opt in field_def.get("options", []):
            opt_label = opt.get("label", "")
            opt_cell = opt.get("cell", "")
            sub_questions = opt.get("sub_questions", [])

            # Container for checkbox + its sub-questions
            opt_container = QWidget()
            opt_layout = QVBoxLayout(opt_container)
            opt_layout.setContentsMargins(0, 0, 0, 0)
            opt_layout.setSpacing(2)

            # The main checkbox for this option
            cb = QCheckBox(opt_label)
            cb.setStyleSheet("font-size: 9pt; font-weight: bold;")
            opt_layout.addWidget(cb)

            # Sub-questions (hidden until checkbox is ticked)
            sub_container: Optional[QWidget] = None
            if sub_questions:
                sub_container = QWidget()
                sub_layout = QFormLayout(sub_container)
                sub_layout.setContentsMargins(24, 2, 0, 4)
                sub_layout.setSpacing(4)
                sub_layout.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
                sub_layout.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)

                for sq in sub_questions:
                    sq_label = sq.get("label", "")
                    sq_cell = sq.get("cell", "")
                    sq_type = sq.get("type", "text").lower()

                    sq_lbl = QLabel(f"{sq_label}:")
                    sq_lbl.setStyleSheet("font-size: 8pt; color: #1a5276;")

                    if sq_type == "multiline":
                        sq_widget: QWidget = QTextEdit()
                        sq_widget.setPlaceholderText(f"Enter {sq_label.lower()} here…")
                        sq_widget.setFixedHeight(60)
                        sq_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                    else:  # "text" (default)
                        sq_widget = QLineEdit()
                        sq_widget.setPlaceholderText(f"Enter {sq_label.lower()} here…")

                    sub_layout.addRow(sq_lbl, sq_widget)
                    if sq_cell:
                        self._sub_widgets[sq_cell] = sq_widget

                sub_container.setVisible(False)
                opt_layout.addWidget(sub_container)

            outer.addWidget(opt_container)

            entry = {
                "cell": opt_cell,
                "checkbox": cb,
                "sub_container": sub_container,
            }
            self._option_checkboxes.append(entry)

            # Connect toggle
            cb.toggled.connect(
                lambda checked, sc=sub_container: sc.setVisible(checked) if sc else None
            )

        self.setLayout(outer)

    # ------------------------------------------------------------------
    # Public helpers
    # ------------------------------------------------------------------

    def get_values(self) -> Dict[str, Any]:
        """Return ``{cell_ref: bool}`` for every option checkbox plus any
        visible sub-question cell values."""
        result: Dict[str, Any] = {}
        for entry in self._option_checkboxes:
            cell = entry["cell"]
            checked: bool = entry["checkbox"].isChecked()
            if cell:
                result[cell] = checked
            # Always read sub-question values so unchecked sub-fields also
            # contribute (they will be empty strings)
        for cell_ref, widget in self._sub_widgets.items():
            if isinstance(widget, QTextEdit):
                result[cell_ref] = widget.toPlainText().strip()
            elif isinstance(widget, QLineEdit):
                result[cell_ref] = widget.text().strip()
        return result

    def clear(self):
        """Uncheck all checkboxes and clear all sub-question fields."""
        for entry in self._option_checkboxes:
            entry["checkbox"].setChecked(False)
            if entry["sub_container"]:
                entry["sub_container"].setVisible(False)
        for widget in self._sub_widgets.values():
            if isinstance(widget, QTextEdit):
                widget.clear()
            elif isinstance(widget, QLineEdit):
                widget.clear()

    def set_values(self, values: Dict[str, Any]):
        """Populate widget from a ``{cell_ref: value}`` mapping (e.g. from a saved form)."""
        self.clear()
        for entry in self._option_checkboxes:
            cell = entry["cell"]
            if cell and cell in values:
                entry["checkbox"].setChecked(bool(values[cell]))
        for cell_ref, widget in self._sub_widgets.items():
            if cell_ref in values:
                val = str(values[cell_ref])
                if isinstance(widget, QTextEdit):
                    widget.setPlainText(val)
                elif isinstance(widget, QLineEdit):
                    widget.setText(val)


# ---------------------------------------------------------------------------
# Sheet-selection dialog (shown before the form)
# ---------------------------------------------------------------------------

class _SheetSelectionDialog(QDialog):
    """Small dialog that lets the user pick which installation sheet to fill."""

    def __init__(self, sheet_names: List[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Installation Sheet")
        self.setMinimumWidth(380)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        layout = QVBoxLayout()
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)

        lbl = QLabel("Please select the installation sheet type:")
        lbl.setStyleSheet("font-size: 10pt; font-weight: bold;")
        layout.addWidget(lbl)

        self.combo = QComboBox()
        self.combo.addItems(sheet_names)
        self.combo.setMinimumHeight(32)
        layout.addWidget(self.combo)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.setLayout(layout)

    @property
    def selected_index(self) -> int:
        return self.combo.currentIndex()


# ---------------------------------------------------------------------------
# Main dialog
# ---------------------------------------------------------------------------

class InstallationSheetDialog(QDialog):
    """Modeless dialog for filling and exporting the installation sheet.

    Opens a sheet-selection dropdown first, then renders a dynamic form whose
    fields (text, multiline, checkbox, auto_fill, or checkbox_group) and Excel
    cell mappings are driven by ``config/installation_sheets.json``.

    The dialog is non-blocking: calling ``show()`` keeps the main application
    fully usable while the sheet is open.
    """

    def __init__(self, parent=None, configuration_name: Optional[str] = None,
                 baudrate: Optional[int] = None, is_offline: bool = False):
        super().__init__(parent)
        self.setWindowTitle("Fill up Installation Sheet")
        self.setMinimumSize(640, 700)
        # Qt.Window makes the dialog modeless and gives it its own taskbar entry;
        # WindowMaximizeButtonHint adds the maximise button.
        self.setWindowFlags(
            Qt.Window
            | Qt.WindowMinimizeButtonHint
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowCloseButtonHint
        )

        self._uploaded_pictures: List[str] = []
        self._field_widgets: Dict[str, QWidget] = {}   # cell_ref -> widget
        self._current_sheet_def: Optional[Dict[str, Any]] = None
        self._configuration_name: Optional[str] = configuration_name
        self._baudrate: Optional[int] = baudrate
        self._is_offline: bool = is_offline

        # Load available sheet definitions
        self._sheets = _load_sheets_config()

        # Ask the user which sheet to use *before* building the form
        if not self._select_sheet():
            # User cancelled selection – close dialog immediately
            self._cancelled = True
            return
        self._cancelled = False
        self._init_ui()

    # ------------------------------------------------------------------
    # Sheet selection
    # ------------------------------------------------------------------

    def _select_sheet(self) -> bool:
        """Show the sheet-selection dialog.  Returns True if accepted."""
        if not self._sheets:
            QMessageBox.warning(
                self,
                "No Sheet Definitions",
                "No installation sheet definitions were found.\n"
                f"Please check: {_sheets_config_path()}",
            )
            return False

        names = [s.get("name", s.get("id", f"Sheet {i}")) for i, s in enumerate(self._sheets)]
        dlg = _SheetSelectionDialog(names, parent=self)
        if dlg.exec_() != QDialog.Accepted:
            return False

        self._current_sheet_def = self._sheets[dlg.selected_index]
        return True

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _init_ui(self):
        sheet_def = self._current_sheet_def
        sheet_name = sheet_def.get("name", "Installation Sheet")
        fields: List[Dict[str, Any]] = sheet_def.get("fields", [])

        outer = QVBoxLayout()
        outer.setSpacing(8)
        outer.setContentsMargins(12, 12, 12, 12)

        # Title label (shows selected sheet name)
        title_lbl = QLabel(sheet_name)
        title_lbl.setStyleSheet(
            "font-size: 15pt; font-weight: bold; color: #1F497D; padding-bottom: 4px;"
        )
        outer.addWidget(title_lbl)

        # ---- Scrollable form area ----
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        form_container = QWidget()
        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        form_layout.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        form_layout.setSpacing(8)
        form_layout.setContentsMargins(4, 4, 4, 4)

        for field in fields:
            label_text = field.get("label", "")
            cell_ref = field.get("cell", "")
            field_type = field.get("type", "text").lower()
            is_required = field.get("required", False)

            if field_type == "section_title":
                # Stand-alone title that groups the questions below it.
                # Rendered as a full-width styled label with no associated widget.
                section_lbl = QLabel(label_text)
                section_lbl.setStyleSheet(
                    "font-size: 10pt; font-weight: bold; color: #1F497D;"
                    " border-bottom: 1px solid #1F497D; padding-bottom: 2px;"
                    " margin-top: 6px;"
                )
                section_lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                form_layout.addRow(section_lbl)
                continue

            # Append " *" to the label for required fields so users can see
            # at a glance which ones must be filled before creating the PDF.
            display_label = f"{label_text} *:" if is_required else f"{label_text}:"
            lbl = QLabel(display_label)
            lbl.setStyleSheet("font-weight: bold; font-size: 9pt;")

            if field_type == "conditional":
                # Conditional question: Yes/No radio buttons + conditional text boxes.
                # The question may be long, so allow the label to word-wrap and
                # align it to the top of the row.
                lbl.setWordWrap(True)
                lbl.setAlignment(Qt.AlignRight | Qt.AlignTop)
                checkbox_cell = field.get("checkbox_cell", "")
                cells_hint = f"Checkbox cell: {checkbox_cell}"
                if field.get("yes_text_cell"):
                    cells_hint += f"  |  Yes text cell: {field['yes_text_cell']}"
                if field.get("no_text_cell"):
                    cells_hint += f"  |  No text cell: {field['no_text_cell']}"
                lbl.setToolTip(cells_hint)
                widget = ConditionalWidget(field)
                # Store using checkbox_cell as key so _read_field_values can
                # detect and call widget.get_values()
                self._field_widgets[checkbox_cell or cell_ref] = widget
                form_layout.addRow(lbl, widget)
                continue

            if field_type == "checkbox_group":
                # Checkbox-group question: multiple named checkboxes each linked to
                # an Excel cell, with optional per-checkbox sub-questions.
                lbl.setWordWrap(True)
                lbl.setAlignment(Qt.AlignRight | Qt.AlignTop)
                widget = CheckboxGroupWidget(field)
                # The group widget manages its own cell→widget mapping; register
                # it under a sentinel key so _read_field_values can find it.
                sentinel = field.get("group_id") or f"__cbg_{len(self._field_widgets)}"
                self._field_widgets[sentinel] = widget
                form_layout.addRow(lbl, widget)
                continue

            # Build a human-readable cell hint for tooltips (range or single cell)
            cell_hint = f"Excel cell{'s' if ':' in cell_ref else ''}: {cell_ref}"
            lbl.setToolTip(cell_hint)

            if field_type == "auto_fill":
                # Auto-fill field: value is populated automatically from context.
                # "auto_source" can be "config_name" or "baudrate".
                auto_source = field.get("auto_source", "")
                if auto_source == "baudrate":
                    # Baudrate: show a dropdown of available CAN baudrates.
                    widget = QComboBox()
                    widget.setMinimumHeight(28)
                    for br in _CAN_BAUDRATES:
                        widget.addItem(str(br), br)
                    # Pre-select the detected baudrate when connected
                    if self._baudrate is not None:
                        idx = widget.findData(self._baudrate)
                        if idx >= 0:
                            widget.setCurrentIndex(idx)
                    widget.setToolTip(f"Select CAN baudrate  |  {cell_hint}")
                elif auto_source == "config_name":
                    # Configuration: show a dropdown of available configurations.
                    config_names = _load_config_names()
                    widget = QComboBox()
                    widget.setMinimumHeight(28)
                    if config_names:
                        widget.addItem("")  # blank/unselected option
                        widget.addItems(config_names)
                    else:
                        widget.addItem("")
                    # Pre-select the current configuration
                    if self._configuration_name:
                        idx = widget.findText(self._configuration_name)
                        if idx >= 0:
                            widget.setCurrentIndex(idx)
                    widget.setToolTip(f"Select configuration  |  {cell_hint}")
                else:
                    auto_value = ""
                    widget = QLineEdit()
                    widget.setPlaceholderText(f"Enter {label_text.lower()} here…")
                    widget.setToolTip(
                        f"Enter manually (not available in current mode)  |  {cell_hint}"
                    )
            elif field_type == "checkbox":
                widget = QCheckBox()
                widget.setToolTip(f"Tick to mark '{label_text}' in {cell_hint}")
            elif field_type == "multiline":
                widget = QTextEdit()
                widget.setPlaceholderText(f"Enter {label_text.lower()} here…")
                widget.setFixedHeight(80)
                widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            else:  # "text"
                widget = QLineEdit()
                widget.setPlaceholderText(f"Enter {label_text.lower()} here…")
                # Pre-populate date fields
                if "date" in label_text.lower():
                    widget.setText(date.today().strftime("%d/%m/%Y"))

            self._field_widgets[cell_ref] = widget
            form_layout.addRow(lbl, widget)

        form_container.setLayout(form_layout)
        scroll.setWidget(form_container)
        outer.addWidget(scroll, stretch=3)

        # ---- Pictures section ----
        pics_group = QGroupBox("Attached Pictures")
        pics_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        pics_layout = QVBoxLayout()
        pics_layout.setSpacing(6)

        self._pics_grid = QGridLayout()
        self._pics_grid.setSpacing(6)

        pics_scroll = QScrollArea()
        pics_scroll.setWidgetResizable(True)
        pics_scroll.setMinimumHeight(110)
        pics_scroll.setMaximumHeight(160)
        pics_scroll.setFrameShape(QFrame.StyledPanel)

        pics_inner = QWidget()
        pics_inner.setLayout(self._pics_grid)
        pics_scroll.setWidget(pics_inner)
        pics_layout.addWidget(pics_scroll)

        self._no_pics_label = QLabel("No pictures added yet.")
        self._no_pics_label.setAlignment(Qt.AlignCenter)
        self._no_pics_label.setStyleSheet("color: #6c757d; font-size: 9pt;")
        self._pics_grid.addWidget(self._no_pics_label, 0, 0)

        pics_group.setLayout(pics_layout)
        outer.addWidget(pics_group, stretch=1)

        # ---- Action buttons ----
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        self._add_pics_btn = QPushButton("📷  Add Pictures")
        self._add_pics_btn.setToolTip("Upload one or more pictures to include in the PDF")
        self._add_pics_btn.clicked.connect(self._add_pictures)
        self._add_pics_btn.setMinimumHeight(36)

        self._save_btn = QPushButton("💾  Save")
        self._save_btn.setToolTip("Save the current form data to a JSON file for later use")
        self._save_btn.clicked.connect(self._save_form)
        self._save_btn.setMinimumHeight(36)

        self._load_btn = QPushButton("📂  Load")
        self._load_btn.setToolTip("Load a previously saved form from a JSON file")
        self._load_btn.clicked.connect(self._load_form)
        self._load_btn.setMinimumHeight(36)

        self._create_pdf_btn = QPushButton("📄  Create PDF")
        self._create_pdf_btn.setToolTip("Fill the Excel template and export it to PDF")
        self._create_pdf_btn.clicked.connect(self._create_pdf)
        self._create_pdf_btn.setMinimumHeight(36)
        self._create_pdf_btn.setStyleSheet(
            "QPushButton { background-color: #1F497D; color: white; font-weight: bold; border-radius: 4px; }"
            "QPushButton:hover { background-color: #2E6DA4; }"
            "QPushButton:pressed { background-color: #163a62; }"
        )

        self._clear_btn = QPushButton("🗑  Clear")
        self._clear_btn.setToolTip("Clear all text boxes and remove pictures")
        self._clear_btn.clicked.connect(self._clear_all)
        self._clear_btn.setMinimumHeight(36)

        btn_layout.addWidget(self._add_pics_btn)
        btn_layout.addWidget(self._save_btn)
        btn_layout.addWidget(self._load_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(self._clear_btn)
        btn_layout.addWidget(self._create_pdf_btn)

        outer.addLayout(btn_layout)
        self.setLayout(outer)

    # Override exec_ so a cancelled selection closes without showing the window
    def exec_(self):
        if getattr(self, "_cancelled", False):
            return QDialog.Rejected
        return super().exec_()

    # ------------------------------------------------------------------
    # Slot: Add Pictures
    # ------------------------------------------------------------------

    def _add_pictures(self):
        """Open a file dialog to select pictures and add thumbnails."""
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Pictures",
            "",
            "Images (*.png *.jpg *.jpeg *.bmp *.gif *.tiff *.webp);;All Files (*)",
        )
        if not paths:
            return

        for p in paths:
            if p not in self._uploaded_pictures:
                self._uploaded_pictures.append(p)

        self._refresh_picture_grid()

    def _refresh_picture_grid(self):
        """Rebuild the thumbnail grid from self._uploaded_pictures."""
        # Clear existing widgets
        while self._pics_grid.count():
            item = self._pics_grid.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        if not self._uploaded_pictures:
            self._no_pics_label = QLabel("No pictures added yet.")
            self._no_pics_label.setAlignment(Qt.AlignCenter)
            self._no_pics_label.setStyleSheet("color: #6c757d; font-size: 9pt;")
            self._pics_grid.addWidget(self._no_pics_label, 0, 0)
            return

        cols = 5
        for idx, path in enumerate(self._uploaded_pictures):
            row, col = divmod(idx, cols)
            thumb_frame = QFrame()
            thumb_frame.setFrameShape(QFrame.Box)
            thumb_frame.setStyleSheet("border: 1px solid #ced4da; border-radius: 4px;")
            thumb_layout = QVBoxLayout()
            thumb_layout.setContentsMargins(2, 2, 2, 2)
            thumb_layout.setSpacing(2)

            # Thumbnail image
            pix = QPixmap(path)
            if not pix.isNull():
                thumb_lbl = QLabel()
                thumb_lbl.setPixmap(pix.scaled(80, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                thumb_lbl.setAlignment(Qt.AlignCenter)
                thumb_layout.addWidget(thumb_lbl)
            else:
                thumb_layout.addWidget(QLabel("(no preview)"))

            # Filename (truncated)
            fname = os.path.basename(path)
            if len(fname) > 12:
                fname = fname[:9] + "…"
            name_lbl = QLabel(fname)
            name_lbl.setAlignment(Qt.AlignCenter)
            name_lbl.setStyleSheet("font-size: 7pt; color: #495057; border: none;")
            thumb_layout.addWidget(name_lbl)

            # Remove button
            remove_btn = QPushButton("✕")
            remove_btn.setFixedSize(20, 20)
            remove_btn.setStyleSheet(
                "QPushButton { background: #dc3545; color: white; border-radius: 2px; "
                "font-size: 8pt; border: none; }"
                "QPushButton:hover { background: #c82333; }"
            )
            _path = path  # capture for lambda
            remove_btn.clicked.connect(lambda _, p=_path: self._remove_picture(p))
            thumb_layout.addWidget(remove_btn, alignment=Qt.AlignCenter)

            thumb_frame.setLayout(thumb_layout)
            self._pics_grid.addWidget(thumb_frame, row, col)

    def _remove_picture(self, path: str):
        """Remove a picture from the list and refresh the grid."""
        if path in self._uploaded_pictures:
            self._uploaded_pictures.remove(path)
        self._refresh_picture_grid()

    # ------------------------------------------------------------------
    # Slot: Clear
    # ------------------------------------------------------------------

    def _clear_all(self):
        """Clear all text fields, uncheck checkboxes, and remove pictures."""
        for widget in self._field_widgets.values():
            if isinstance(widget, QLineEdit):
                if not widget.isReadOnly():
                    widget.clear()
            elif isinstance(widget, QTextEdit):
                widget.clear()
            elif isinstance(widget, QCheckBox):
                widget.setChecked(False)
            elif isinstance(widget, ConditionalWidget):
                widget.clear()
            elif isinstance(widget, CheckboxGroupWidget):
                widget.clear()
        self._uploaded_pictures.clear()
        self._refresh_picture_grid()

    # ------------------------------------------------------------------
    # Slot: Save / Load form
    # ------------------------------------------------------------------

    def _save_form(self):
        """Save the current form state (field values + picture paths) to a JSON file."""
        sheet_def = self._current_sheet_def or {}
        default_name = f"installation_{sheet_def.get('id', 'sheet')}.json"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Installation Sheet", default_name,
            "Installation Sheet Files (*.json);;All Files (*)"
        )
        if not save_path:
            return

        data = {
            "sheet_id": sheet_def.get("id", ""),
            "field_values": self._read_field_values(),
            "pictures": list(self._uploaded_pictures),
        }
        try:
            with open(save_path, "w", encoding="utf-8") as fh:
                json.dump(data, fh, indent=2, ensure_ascii=False)
            QMessageBox.information(
                self, "Saved",
                f"Form data saved successfully:\n{save_path}",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Save Error", f"Could not save form:\n{exc}")

    def _load_form(self):
        """Load a previously saved form JSON and populate all widgets."""
        load_path, _ = QFileDialog.getOpenFileName(
            self, "Load Installation Sheet", "",
            "Installation Sheet Files (*.json);;All Files (*)"
        )
        if not load_path:
            return

        try:
            with open(load_path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
        except Exception as exc:
            QMessageBox.critical(self, "Load Error", f"Could not load form:\n{exc}")
            return

        field_values: Dict[str, Any] = data.get("field_values", {})
        pictures: List[str] = data.get("pictures", [])

        # Restore widget values
        for cell_ref, widget in self._field_widgets.items():
            if isinstance(widget, ConditionalWidget):
                widget.set_values(field_values)
            elif isinstance(widget, CheckboxGroupWidget):
                widget.set_values(field_values)
            elif isinstance(widget, QComboBox):
                val = field_values.get(cell_ref, "")
                idx = widget.findText(str(val))
                if idx >= 0:
                    widget.setCurrentIndex(idx)
            elif isinstance(widget, QLineEdit):
                if not widget.isReadOnly():
                    widget.setText(str(field_values.get(cell_ref, "")))
            elif isinstance(widget, QTextEdit):
                widget.setPlainText(str(field_values.get(cell_ref, "")))
            elif isinstance(widget, QCheckBox):
                widget.setChecked(bool(field_values.get(cell_ref, False)))

        # Restore pictures (only keep paths that still exist on disk)
        self._uploaded_pictures = [p for p in pictures if os.path.isfile(p)]
        missing = [p for p in pictures if not os.path.isfile(p)]
        self._refresh_picture_grid()

        if missing:
            QMessageBox.warning(
                self, "Missing Pictures",
                "The following picture files could not be found and were skipped:\n"
                + "\n".join(missing),
            )

    # ------------------------------------------------------------------
    # Slot: Create PDF
    # ------------------------------------------------------------------

    def _validate_required_fields(self) -> List[str]:
        """Return a list of labels for required fields that have not been filled.

        Checks:
        - ``text`` / ``multiline`` / ``auto_fill``: the mapped cell value must be
          a non-empty string.
        - ``conditional``: at least one of the Yes / No radio buttons must be
          selected (i.e. ``checkbox_cell`` must be present in the collected values).

        ``checkbox``, ``checkbox_group``, and ``section_title`` fields are always
        considered valid because checkboxes inherently carry a value.
        """
        missing: List[str] = []
        sheet_def = self._current_sheet_def or {}
        fields: List[Dict[str, Any]] = sheet_def.get("fields", [])
        field_values = self._read_field_values()

        for field in fields:
            if not field.get("required", False):
                continue
            label_text = field.get("label", "")
            field_type = field.get("type", "text").lower()

            if field_type in ("text", "multiline", "auto_fill"):
                cell_ref = field.get("cell", "")
                val = field_values.get(cell_ref, "")
                if not str(val).strip():
                    missing.append(label_text)
            elif field_type == "conditional":
                # The checkbox_cell key is only present when a radio button is
                # selected; its absence means neither Yes nor No was chosen.
                checkbox_cell = field.get("checkbox_cell", "")
                if checkbox_cell not in field_values:
                    missing.append(label_text)

        return missing

    def _create_pdf(self):
        """
        Fill the Excel template with form data, embed pictures, then export to PDF.

        Strategy:
          1. Copy the template to a temporary file.
          2. Write each field value to its mapped cell using openpyxl.
          3. Insert uploaded pictures below the data rows.
          4. Try to export via win32com (Excel on Windows).
          5. Fall back to reportlab if win32com is unavailable.
        """
        # --- Validate required fields before anything else ---
        missing = self._validate_required_fields()
        if missing:
            QMessageBox.warning(
                self,
                "Required Fields Missing",
                "Please fill in the following required fields before creating the PDF:\n\n"
                + "\n".join(f"  \u2022 {m}" for m in missing),
            )
            return

        # --- Ask where to save the PDF ---
        default_name = "installation_sheet.pdf"
        pdf_path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF", default_name, "PDF Files (*.pdf)"
        )
        if not pdf_path:
            return

        try:
            self._do_create_pdf(pdf_path)
            QMessageBox.information(
                self,
                "PDF Created",
                f"Installation sheet saved successfully:\n{pdf_path}",
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Error Creating PDF",
                f"Could not create PDF:\n{exc}",
            )

    def _read_field_values(self) -> Dict[str, Any]:
        """Return a mapping of excel_cell → current value (str or bool).

        For :class:`ConditionalWidget` entries the widget's own
        :meth:`~ConditionalWidget.get_values` method is called, which may
        contribute up to three cell mappings (checkbox_cell, yes_text_cell,
        no_text_cell).

        For :class:`CheckboxGroupWidget` entries the widget's own
        :meth:`~CheckboxGroupWidget.get_values` method is called, which
        contributes one bool per checkbox cell plus any sub-question cells.
        """
        values: Dict[str, Any] = {}
        for cell_ref, widget in self._field_widgets.items():
            if isinstance(widget, ConditionalWidget):
                values.update(widget.get_values())
            elif isinstance(widget, CheckboxGroupWidget):
                values.update(widget.get_values())
            elif isinstance(widget, QComboBox):
                values[cell_ref] = widget.currentText().strip()
            elif isinstance(widget, QLineEdit):
                values[cell_ref] = widget.text().strip()
            elif isinstance(widget, QTextEdit):
                values[cell_ref] = widget.toPlainText().strip()
            elif isinstance(widget, QCheckBox):
                values[cell_ref] = widget.isChecked()
        return values

    def _embed_pictures_in_excel(self, ws, anchor_cell: str = None):
        """Embed uploaded pictures directly into the Excel worksheet.

        Inserts each picture as an over-cell image (Excel's "Insert Picture
        Over Cells") starting at *anchor_cell* (e.g. ``"A35"``).  Every image
        is unconditionally resized to exactly 25 cm wide × 45 cm tall so the
        output is consistent regardless of the original picture dimensions.
        Note: images that do not share the 25:45 aspect ratio will appear
        stretched or compressed; this is intentional to ensure a uniform size
        in the exported PDF.  When no anchor is given the pictures are placed
        three rows below the last populated row.  Multiple pictures are stacked
        vertically, each starting directly below the previous one.
        """
        if not self._uploaded_pictures:
            return

        try:
            from openpyxl.drawing.image import Image as OXLImage
        except ImportError:
            return  # openpyxl image support unavailable – skip

        # Determine the starting row / column for picture insertion
        if anchor_cell:
            m = re.match(r"([A-Za-z]+)(\d+)$", anchor_cell.strip())
            if m:
                start_col_letter = m.group(1).upper()
                current_row = int(m.group(2))
            else:
                start_col_letter = "A"
                current_row = (ws.max_row or 0) + 3
        else:
            start_col_letter = "A"
            current_row = (ws.max_row or 0) + 3

        # Fixed target size: 25 cm wide × 45 cm tall at 96 DPI (1 cm ≈ 37.795 px)
        CM_TO_PX = 96 / 2.54
        target_w_px = round(25 * CM_TO_PX)   # ≈ 945 px
        target_h_px = round(45 * CM_TO_PX)   # ≈ 1701 px
        # Approximate Excel row height in pixels (default row ≈ 15 pt ≈ 20 px at 96 DPI)
        px_per_row = 20

        for pic_path in self._uploaded_pictures:
            if not os.path.isfile(pic_path):
                continue
            try:
                img = OXLImage(pic_path)
                # Always force the picture to exactly 25 cm × 45 cm
                img.width = target_w_px
                img.height = target_h_px

                img.anchor = f"{start_col_letter}{current_row}"
                ws.add_image(img)

                # Advance the row pointer so the next picture starts below this one
                rows_used = max(1, target_h_px // px_per_row) + 2
                current_row += rows_used
            except (IOError, OSError, ValueError):
                continue  # skip unreadable or unsupported images

    def _do_create_pdf(self, pdf_path: str):
        """Internal: fill Excel template from config/, embed pictures, export to PDF.

        Strategy
        --------
        Attempt 1 – win32com / Excel (Windows):
          1. Copy the template to a temporary file.
          2. Open the copy **directly** with Excel via win32com – no openpyxl
             round-trip.  This preserves every native Excel feature: Form
             Control checkboxes, coloured shapes/drawings, and cell formatting.
          3. Write each field value into its mapped cell via ``Range.Value``.
             Boolean checkbox values are written as ☑ / ☐ characters so that
             unchecked checkboxes never appear as "FALSE" in the exported PDF.
          4. Insert uploaded pictures via Excel's ``Shapes.AddPicture`` API.
          5. Export to PDF with ``ExportAsFixedFormat``.
          6. Close the temporary workbook without saving.

        Attempt 2 – reportlab fallback (no Excel / win32com available):
          Render a simple formatted PDF directly from the collected field values
          using reportlab.  openpyxl is not required for this path.
        """
        sheet_def = self._current_sheet_def or {}
        excel_filename = sheet_def.get("excel_file", "")
        sheet_index = sheet_def.get("sheet_index", 0)
        picture_anchor = sheet_def.get("picture_anchor_cell", None)

        # --- Locate the Excel template in config/ ---
        if excel_filename:
            tpl = _excel_template_path(excel_filename)
        else:
            tpl = ""

        if not tpl or not os.path.isfile(tpl):
            raise RuntimeError(
                f"Excel template not found in config folder: {excel_filename or '(none specified)'}\n"
                f"Please place the file in: {_config_dir()}"
            )

        # Collect field values from the Qt form widgets now so both code paths
        # can use them without repeating the work.
        field_values = self._read_field_values()

        # Checkbox symbols can be overridden per-sheet in installation_sheets.json.
        checkbox_yes_char = sheet_def.get("checkbox_yes_char", "☑")
        checkbox_no_char = sheet_def.get("checkbox_no_char", "☐")
        fields: List[Dict[str, Any]] = sheet_def.get("fields", [])

        pdf_created = False

        # ------------------------------------------------------------------
        # Attempt 1: win32com – open the template copy directly with Excel.
        # Bypassing the openpyxl load/save cycle preserves Form Controls,
        # coloured shapes, and all cell formatting exactly as designed.
        # ------------------------------------------------------------------
        if not pdf_created:
            try:
                import win32com.client
                import pythoncom

                tmp_dir = tempfile.mkdtemp(prefix="sartel_install_")
                tmp_xlsx = os.path.join(tmp_dir, "installation_sheet_filled.xlsx")
                shutil.copy2(tpl, tmp_xlsx)

                try:
                    pythoncom.CoInitialize()
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    try:
                        wb_com = excel.Workbooks.Open(os.path.abspath(tmp_xlsx))

                        # COM worksheet indices are 1-based.
                        if sheet_index < wb_com.Worksheets.Count:
                            ws_com = wb_com.Worksheets(sheet_index + 1)
                        else:
                            ws_com = wb_com.ActiveSheet

                        # Before writing user values, pre-fill every checkbox
                        # cell defined in the JSON that the user left untouched
                        # with the unchecked symbol.  This ensures that any
                        # cell that was displaying "FALSE" from a native Form
                        # Control link will show ☐ instead.
                        for fld in fields:
                            ftype = fld.get("type", "text").lower()
                            if ftype == "checkbox":
                                c = fld.get("cell", "")
                                if c and c not in field_values:
                                    ws_com.Range(_com_cell_ref(c)).Value = checkbox_no_char
                            elif ftype == "conditional":
                                c = fld.get("checkbox_cell", "")
                                if c and c not in field_values:
                                    ws_com.Range(_com_cell_ref(c)).Value = checkbox_no_char
                            elif ftype == "checkbox_group":
                                for opt in fld.get("options", []):
                                    c = opt.get("cell", "")
                                    if c and c not in field_values:
                                        ws_com.Range(_com_cell_ref(c)).Value = checkbox_no_char

                        # Write user-provided field values.
                        for cell_ref, value in field_values.items():
                            com_ref = _com_cell_ref(cell_ref)
                            if isinstance(value, bool):
                                ws_com.Range(com_ref).Value = (
                                    checkbox_yes_char if value else checkbox_no_char
                                )
                            else:
                                ws_com.Range(com_ref).Value = value

                        # Insert uploaded pictures via Excel's Shapes API so
                        # they appear at the correct position in the exported PDF.
                        if self._uploaded_pictures:
                            # Target picture size: 25 cm wide × 45 cm tall.
                            # Excel COM measures positions/sizes in points
                            # (1 pt = 1/72 inch; 1 cm ≈ 28.3465 pt).
                            CM_TO_PT = 72.0 / 2.54
                            pic_w = round(25 * CM_TO_PT)
                            pic_h = round(45 * CM_TO_PT)

                            # Starting position: the anchor cell when provided,
                            # otherwise top-left of the used range plus a gap.
                            if picture_anchor:
                                try:
                                    anc = ws_com.Range(_com_cell_ref(picture_anchor))
                                    left_pos = float(anc.Left)
                                    top_pos = float(anc.Top)
                                except Exception:
                                    left_pos, top_pos = 0.0, 0.0
                            else:
                                left_pos, top_pos = 0.0, 0.0

                            for pic_path in self._uploaded_pictures:
                                if not os.path.isfile(pic_path):
                                    continue
                                try:
                                    ws_com.Shapes.AddPicture(
                                        os.path.abspath(pic_path),
                                        False,   # LinkToFile
                                        True,    # SaveWithDocument
                                        left_pos,
                                        top_pos,
                                        pic_w,
                                        pic_h,
                                    )
                                    # Stack pictures vertically with a small gap.
                                    top_pos += pic_h + 10
                                except Exception:
                                    pass  # skip unreadable / unsupported images

                        wb_com.ExportAsFixedFormat(
                            0,      # xlTypePDF
                            os.path.abspath(pdf_path),
                            1,      # xlQualityStandard
                            True,
                            False,
                        )
                        wb_com.Close(False)
                        pdf_created = True
                    finally:
                        excel.Quit()
                        pythoncom.CoUninitialize()
                finally:
                    try:
                        shutil.rmtree(tmp_dir, ignore_errors=True)
                    except Exception:
                        pass
            except Exception:
                # A broad catch is intentional: win32com raises pywintypes.com_error
                # (a dynamic type only importable when pywin32 is installed) in
                # addition to ImportError when the package is absent.  Any failure
                # here falls through to the reportlab fallback below.
                pass

        # ------------------------------------------------------------------
        # Attempt 2: reportlab fallback (no Excel / win32com available).
        # ------------------------------------------------------------------
        if not pdf_created:
            self._create_pdf_reportlab(pdf_path, field_values)
            pdf_created = True  # noqa: F841

    def _create_pdf_reportlab(self, pdf_path: str, field_values: Dict[str, Any]):
        """Generate a formatted PDF using reportlab when Excel COM is not available."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                Image as RLImage, HRFlowable,
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError as exc:
            raise RuntimeError(
                "Neither Microsoft Excel (win32com) nor the 'reportlab' package is available. "
                "Install reportlab with:  pip install reportlab"
            ) from exc

        styles = getSampleStyleSheet()
        sheet_def = self._current_sheet_def or {}
        sheet_name = sheet_def.get("name", "Installation Sheet")
        fields: List[Dict[str, Any]] = sheet_def.get("fields", [])

        title_style = ParagraphStyle(
            "SheetTitle",
            parent=styles["Title"],
            fontSize=16,
            textColor=colors.HexColor("#1F497D"),
            spaceAfter=6,
            alignment=TA_CENTER,
        )
        sub_style = ParagraphStyle(
            "SheetSub",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#4472C4"),
            spaceAfter=12,
            alignment=TA_CENTER,
            italic=True,
        )
        field_label_style = ParagraphStyle(
            "FieldLabel",
            parent=styles["Normal"],
            fontSize=9,
            fontName="Helvetica-Bold",
        )
        field_value_style = ParagraphStyle(
            "FieldValue",
            parent=styles["Normal"],
            fontSize=9,
        )
        section_title_style = ParagraphStyle(
            "SectionTitle",
            parent=styles["Normal"],
            fontSize=9,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1F497D"),
        )

        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            leftMargin=20 * mm,
            rightMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )
        story = []

        story.append(Paragraph(f"SARTEL – {sheet_name}", title_style))
        story.append(Paragraph("CAN Bus Monitoring – Field Installation Record", sub_style))
        story.append(HRFlowable(width="100%", color=colors.HexColor("#4472C4"), thickness=1.5))
        story.append(Spacer(1, 8 * mm))

        # Build table rows from the dynamic field list.
        # section_title rows span both columns; track their row indices so the
        # TableStyle can apply SPAN and a distinct background colour.
        table_data = [
            [
                Paragraph("<b>Field</b>", field_label_style),
                Paragraph("<b>Value</b>", field_label_style),
            ]
        ]
        section_title_rows: List[int] = []
        for field in fields:
            label_text = field.get("label", "")
            field_type = field.get("type", "text").lower()

            if field_type == "section_title":
                # Render as a full-width header row that spans both columns.
                row_idx = len(table_data)
                section_title_rows.append(row_idx)
                table_data.append([
                    Paragraph(label_text, section_title_style),
                    Paragraph("", field_value_style),
                ])
                continue

            if field_type == "conditional":
                # Show the boolean answer + up to two associated text notes
                checkbox_cell = field.get("checkbox_cell", "")
                yes_text_cell = field.get("yes_text_cell", "")
                yes_text_cell2 = field.get("yes_text_cell2", "")
                no_text_cell = field.get("no_text_cell", "")
                no_text_cell2 = field.get("no_text_cell2", "")
                raw_bool = field_values.get(checkbox_cell)
                if raw_bool is True:
                    parts = ["☑ Yes"]
                    if field_values.get(yes_text_cell):
                        parts.append(field_values[yes_text_cell])
                    if field_values.get(yes_text_cell2):
                        parts.append(field_values[yes_text_cell2])
                    display_val = " – ".join(parts)
                elif raw_bool is False:
                    parts = ["☐ No"]
                    if field_values.get(no_text_cell):
                        parts.append(field_values[no_text_cell])
                    if field_values.get(no_text_cell2):
                        parts.append(field_values[no_text_cell2])
                    display_val = " – ".join(parts)
                else:
                    display_val = "—"
            elif field_type == "checkbox_group":
                # Show each option as its own sub-row
                options = field.get("options", [])
                parts = []
                for opt in options:
                    opt_label = opt.get("label", "")
                    opt_cell = opt.get("cell", "")
                    checked = field_values.get(opt_cell, False)
                    icon = "☑" if checked else "☐"
                    sub_parts = [f"{icon} {opt_label}"]
                    if checked:
                        for sq in opt.get("sub_questions", []):
                            sq_cell = sq.get("cell", "")
                            sq_val = field_values.get(sq_cell, "")
                            if sq_val:
                                sub_parts.append(f"  {sq.get('label','')}: {sq_val}")
                    parts.append("  ".join(sub_parts))
                display_val = "\n".join(parts) if parts else "—"
            else:
                cell_ref = field.get("cell", "")
                raw_val = field_values.get(cell_ref, "")
                # Format booleans as checkbox icons
                if isinstance(raw_val, bool):
                    display_val = "☑ Yes" if raw_val else "☐ No"
                else:
                    display_val = str(raw_val) if raw_val else "—"

            table_data.append([
                Paragraph(label_text, field_label_style),
                Paragraph(display_val, field_value_style),
            ])

        page_w = A4[0] - 40 * mm
        col_widths = [page_w * 0.38, page_w * 0.62]

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        header_bg = colors.HexColor("#4472C4")
        section_bg = colors.HexColor("#D6E4F0")
        even_bg = colors.HexColor("#DCE6F1")
        white = colors.white

        tbl_style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR",  (0, 0), (-1, 0), white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, even_bg]),
            ("GRID",  (0, 0), (-1, -1), 0.5, colors.HexColor("#ADB5BD")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ]
        # Merge both columns for every section-title row and apply a distinct
        # background colour so the heading stands out visually.
        for row_idx in section_title_rows:
            tbl_style_cmds.append(("SPAN", (0, row_idx), (1, row_idx)))
            tbl_style_cmds.append(("BACKGROUND", (0, row_idx), (1, row_idx), section_bg))
            tbl_style_cmds.append(("TEXTCOLOR", (0, row_idx), (1, row_idx), colors.HexColor("#1F497D")))

        tbl_style = TableStyle(tbl_style_cmds)
        tbl.setStyle(tbl_style)
        story.append(tbl)

        # Attached pictures
        if self._uploaded_pictures:
            story.append(Spacer(1, 8 * mm))
            story.append(HRFlowable(width="100%", color=colors.HexColor("#4472C4"), thickness=1))
            story.append(Spacer(1, 4 * mm))
            story.append(Paragraph("Attached Pictures", title_style))
            story.append(Spacer(1, 4 * mm))

            max_w = page_w * 0.75
            max_h = 120 * mm

            for pic_path in self._uploaded_pictures:
                if not os.path.isfile(pic_path):
                    continue
                try:
                    img = RLImage(pic_path)
                    # Scale proportionally to fit
                    scale = min(max_w / img.imageWidth, max_h / img.imageHeight, 1.0)
                    img.drawWidth  = img.imageWidth  * scale
                    img.drawHeight = img.imageHeight * scale
                    story.append(img)
                    story.append(Spacer(1, 4 * mm))
                    fname = os.path.basename(pic_path)
                    story.append(
                        Paragraph(
                            fname,
                            ParagraphStyle(
                                "PicCaption", parent=styles["Normal"],
                                fontSize=8, textColor=colors.grey, alignment=TA_CENTER,
                            ),
                        )
                    )
                    story.append(Spacer(1, 6 * mm))
                except Exception:
                    pass  # skip unreadable images

        doc.build(story)
